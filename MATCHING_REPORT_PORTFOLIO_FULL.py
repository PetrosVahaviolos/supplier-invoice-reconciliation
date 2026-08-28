"""
Supplier Invoice Reconciliation Engine
=======================================
Author: Petros Vachaviolos
Description:
    Matches unposted supplier invoices against internal SAP records.
    Categorizes each entry as: exact match, wrong amount, wrong date,
    wrong reference, or unmatched.

Output:
    MATCHING_RESULT.xlsx with 5 sheets:
        ΒΡΕΘΗΚΑΝ          - Exact matches
        ΛΑΘΟΣ ΑΝΑΦΟΡΑ     - Wrong reference
        ΔΕΝ ΒΡΕΘΗΚΑΝ      - No match found
        ΛΑΘΟΣ ΗΜΕΡΟΜΗΝΙΑ  - Wrong date
        ΛΑΘΟΣ ΑΞΙΑ        - Wrong amount
"""

import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PATH = "./"

# =============================================================================
# PHASE 1 — DUMMY DATA GENERATION (for demo/portfolio purposes only)
#
# In production, this phase is fully replaced by automated data collection:
#   ΕΛΒΑΛ.xlsx        → SAP transaction export via Power Automate Desktop
#   ΠΡΟΜΗΘΕΥΤΕΣ.xlsx  → SAP supplier master export via Power Automate Desktop
#   ΑΚΑΤΑΧΩΡΗΤΑ.xlsx  → Unposted entries file shared by the finance team
#                        (ELVALHALCOR + HALCOR combined unposted invoices)
#
# The automation reads directly from SAP GUI using pywinauto and saves the
# exports to a local folder, eliminating manual data collection entirely.
# =============================================================================

random.seed(42)

SUPPLIERS = [
    {"NAME": "ALPHA INDUSTRIES SA",  "VAT": "123456789"},
    {"NAME": "BETA TRADING SA",       "VAT": "987654321"},
    {"NAME": "GAMMA SUPPLIES SA",     "VAT": "456789123"},
    {"NAME": "DELTA METALS SA",       "VAT": "321654987"},
    {"NAME": "EPSILON LOGISTICS SA",  "VAT": "654321098"},
]

def random_date(start="2024-01-01", end="2024-12-31"):
    start_dt = datetime.strptime(start, "%Y-%m-%d")
    end_dt   = datetime.strptime(end,   "%Y-%m-%d")
    return (start_dt + timedelta(days=random.randint(0, (end_dt - start_dt).days))).strftime("%Y-%m-%d")

def random_ref(prefix="TIM", year=2024):
    return f"{prefix}-{year}-{random.randint(1000, 9999)}"

# Supplier master
df_prom_raw = pd.DataFrame(SUPPLIERS)
df_prom_raw["PAYMENT_TERMS"] = random.choices(["30 days", "60 days", "90 days"], k=len(SUPPLIERS))
df_prom_raw["COUNTRY"]       = random.choices(["GR", "DE", "IT", "FR"], k=len(SUPPLIERS))

# SAP transactions (ΕΛΒΑΛ)
sap_rows = []
for i in range(80):
    sup  = random.choice(SUPPLIERS)
    sap_rows.append({
        "DOC_ID":  f"SAP{i+1:04d}",
        "VAT":     sup["VAT"],
        "REF":     random_ref(),
        "AMOUNT":  round(random.uniform(500, 15000), 2),
        "DATE":    random_date(),
        "STATUS":  random.choice(["POSTED", "POSTED", "POSTED", "PENDING"]),
    })
df_elval_raw = pd.DataFrame(sap_rows)

# Unposted entries (ΑΚΑΤΑΧΩΡΗΤΑ) — mix of match scenarios
akat_rows = []

# 30 exact matches
for row in random.sample(sap_rows, 30):
    num = row["REF"].split("-")[-1]
    akat_rows.append({"VAT": row["VAT"], "REF": num, "AMOUNT": row["AMOUNT"], "DATE": row["DATE"]})

# 15 wrong amount
for row in random.sample(sap_rows, 15):
    num = row["REF"].split("-")[-1]
    akat_rows.append({"VAT": row["VAT"], "REF": num, "AMOUNT": round(row["AMOUNT"] * random.uniform(0.8, 1.2), 2), "DATE": row["DATE"]})

# 10 wrong date
for row in random.sample(sap_rows, 10):
    num = row["REF"].split("-")[-1]
    akat_rows.append({"VAT": row["VAT"], "REF": num, "AMOUNT": row["AMOUNT"], "DATE": random_date()})

# 10 wrong reference
for row in random.sample(sap_rows, 10):
    akat_rows.append({"VAT": row["VAT"], "REF": str(random.randint(5000, 9999)), "AMOUNT": row["AMOUNT"], "DATE": row["DATE"]})

# 15 unmatched (ΧΑΛΚΟΡ entries — different company, no match expected)
for _ in range(15):
    sup = random.choice(SUPPLIERS)
    akat_rows.append({"VAT": sup["VAT"], "REF": str(random.randint(1000, 4999)), "AMOUNT": round(random.uniform(500, 15000), 2), "DATE": random_date()})

df_akat_raw = pd.DataFrame(akat_rows)

# Save dummy files
df_prom_raw.to_excel(PATH + "ΠΡΟΜΗΘΕΥΤΕΣ.xlsx", index=False)
df_elval_raw.to_excel(PATH + "ΕΛΒΑΛ.xlsx", index=False)
df_akat_raw.to_excel(PATH + "ΑΚΑΤΑΧΩΡΗΤΑ.xlsx", index=False)

print("Phase 1 complete — dummy data generated.")
print(f"  ΠΡΟΜΗΘΕΥΤΕΣ: {len(df_prom_raw)} suppliers")
print(f"  ΕΛΒΑΛ:       {len(df_elval_raw)} SAP transactions")
print(f"  ΑΚΑΤΑΧΩΡΗΤΑ: {len(df_akat_raw)} unposted entries")
print()

# =============================================================================
# PHASE 2 — MATCHING REPORT ENGINE
#
# Loads the 3 input files, joins them on VAT, and categorizes each
# unposted entry based on how well it matches the SAP records.
#
# Matching logic:
#   EXACT          : VAT + DATE + AMOUNT + REF (contains) all match
#   ΛΑΘΟΣ ΑΞΙΑ    : VAT + DATE + REF match but AMOUNT differs
#   ΛΑΘΟΣ ΗΜΕΡ.   : VAT + AMOUNT + REF match but DATE differs
#   ΛΑΘΟΣ ΑΝΑΦΟΡΑ : SAP match found but REF contains check fails
#   ΔΕΝ ΒΡΕΘΗΚΑΝ  : No SAP record found at all
#
# Edge case: supplier may send "3" as reference which could falsely
# match "13" or "31" — flagged separately as "ΠΡΟΣ ΕΞΕΤΑΣΗ".
# =============================================================================

# Load files
df_prom = pd.read_excel(PATH + "ΠΡΟΜΗΘΕΥΤΕΣ.xlsx")
df_prom["VAT"] = df_prom["VAT"].astype(str).str.strip()

df_elval = pd.read_excel(PATH + "ΕΛΒΑΛ.xlsx")
df_elval["DATE"]   = pd.to_datetime(df_elval["DATE"], errors="coerce")
df_elval["AMOUNT"] = pd.to_numeric(df_elval["AMOUNT"], errors="coerce")
df_elval["VAT"]    = df_elval["VAT"].astype(str).str.strip()

# Extract numeric part from REF (e.g. "TIM-2024-1234" → 1234)
def extract_number(txt):
    txt = str(txt)
    for i in range(len(txt) - 1, -1, -1):
        if not txt[i].isdigit():
            return txt[i + 1:]
    return txt

df_elval["REF_NUM"] = df_elval["REF"].apply(extract_number)
df_elval["REF_NUM"] = pd.to_numeric(df_elval["REF_NUM"], errors="coerce").astype("Int64")

# Rename SAP columns with SAP_ prefix to avoid column conflicts on merge
df_sap = df_elval.rename(columns={
    "VAT": "SAP_VAT", "REF": "SAP_REF", "REF_NUM": "SAP_REF_NUM",
    "DOC_ID": "SAP_DOC_ID", "AMOUNT": "SAP_AMOUNT", "DATE": "SAP_DATE",
})
SAP_COLS = ["SAP_VAT", "SAP_REF", "SAP_REF_NUM", "SAP_DOC_ID", "SAP_AMOUNT", "SAP_DATE"]

df_akat = pd.read_excel(PATH + "ΑΚΑΤΑΧΩΡΗΤΑ.xlsx")
df_akat["DATE"]   = pd.to_datetime(df_akat["DATE"], errors="coerce")
df_akat["AMOUNT"] = pd.to_numeric(df_akat["AMOUNT"], errors="coerce")
df_akat["REF"]    = pd.to_numeric(df_akat["REF"], errors="coerce").astype("Int64")
df_akat["VAT"]    = df_akat["VAT"].astype(str).str.strip()
df_akat = df_akat.merge(df_prom[["VAT", "NAME"]], on="VAT", how="left")

# Matching functions
EMPTY_SAP = {col: None for col in SAP_COLS}

def find_exact(row):
    ref_str = str(row["REF"])
    return df_sap[
        (df_sap["SAP_VAT"] == row["VAT"]) &
        (df_sap["SAP_DATE"] == row["DATE"]) &
        ((df_sap["SAP_AMOUNT"] - row["AMOUNT"]).abs() < 0.01) &
        (df_sap["SAP_REF_NUM"].astype(str).apply(lambda x: x in ref_str))
    ][SAP_COLS].reset_index(drop=True)

def find_wrong_date(row):
    ref_str = str(row["REF"])
    return df_sap[
        (df_sap["SAP_VAT"] == row["VAT"]) &
        (df_sap["SAP_DATE"] != row["DATE"]) &
        ((df_sap["SAP_AMOUNT"] - row["AMOUNT"]).abs() < 0.01) &
        (df_sap["SAP_REF_NUM"].astype(str).apply(lambda x: x in ref_str))
    ][SAP_COLS].reset_index(drop=True)

def find_wrong_amount(row):
    ref_str = str(row["REF"])
    return df_sap[
        (df_sap["SAP_VAT"] == row["VAT"]) &
        (df_sap["SAP_DATE"] == row["DATE"]) &
        ((df_sap["SAP_AMOUNT"] - row["AMOUNT"]).abs() >= 0.01) &
        (df_sap["SAP_REF_NUM"].astype(str).apply(lambda x: x in ref_str))
    ][SAP_COLS].reset_index(drop=True)

# Run matching loop
results        = []
results_date   = []
results_amount = []

for _, row in df_akat.iterrows():
    akat_dict = row.to_dict()
    matched = find_exact(row)
    if len(matched) == 0:
        results.append({**akat_dict, **EMPTY_SAP})
        for _, m in find_wrong_date(row).iterrows():
            results_date.append({**akat_dict, **m.to_dict()})
        for _, m in find_wrong_amount(row).iterrows():
            results_amount.append({**akat_dict, **m.to_dict()})
    else:
        for _, m in matched.iterrows():
            results.append({**akat_dict, **m.to_dict()})

# Build output DataFrames
OUTPUT_COLS = ["NAME", "VAT", "REF", "AMOUNT", "DATE",
               "SAP_DOC_ID", "SAP_REF", "SAP_REF_NUM", "SAP_AMOUNT", "SAP_DATE"]

def build_df(rows):
    if not rows:
        return pd.DataFrame(columns=OUTPUT_COLS)
    df = pd.DataFrame(rows)
    if "SAP_REF_NUM" in df.columns:
        df["SAP_REF_NUM"] = pd.to_numeric(df["SAP_REF_NUM"], errors="coerce").astype("Int64")
    cols = [c for c in OUTPUT_COLS if c in df.columns]
    return df[cols]

def format_dates(df):
    for col in df.columns:
        if "DATE" in col.upper():
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d/%m/%Y")
    return df

def to_int_str(series):
    return series.apply(lambda x: str(int(x)) if pd.notna(x) else "")

df_result = build_df(results)
df_date   = build_df(results_date)
df_amount = build_df(results_amount)

# Split into 5 output sheets
df_found     = df_result[df_result["SAP_DOC_ID"].notna() & (to_int_str(df_result["REF"]) == to_int_str(df_result["SAP_REF_NUM"]))]
df_wrong_ref = df_result[df_result["SAP_DOC_ID"].notna() & (to_int_str(df_result["REF"]) != to_int_str(df_result["SAP_REF_NUM"]))]
df_not_found = df_result[df_result["SAP_DOC_ID"].isna()]

df_found     = format_dates(df_found.copy())
df_wrong_ref = format_dates(df_wrong_ref.copy())
df_not_found = format_dates(df_not_found.copy())
df_date      = format_dates(df_date.copy())
df_amount    = format_dates(df_amount.copy())

# Write to Excel
OUTPUT = PATH + "MATCHING_RESULT.xlsx"

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    df_found.to_excel(writer,     sheet_name="ΒΡΕΘΗΚΑΝ",          index=False)
    df_wrong_ref.to_excel(writer, sheet_name="ΛΑΘΟΣ ΑΝΑΦΟΡΑ",     index=False)
    df_not_found.to_excel(writer, sheet_name="ΔΕΝ ΒΡΕΘΗΚΑΝ",      index=False)
    df_date.to_excel(writer,      sheet_name="ΛΑΘΟΣ ΗΜΕΡΟΜΗΝΙΑ",  index=False)
    df_amount.to_excel(writer,    sheet_name="ΛΑΘΟΣ ΑΞΙΑ",         index=False)

# Autofit columns + table style
wb = load_workbook(OUTPUT)
for sheet_name, df in [
    ("ΒΡΕΘΗΚΑΝ", df_found), ("ΛΑΘΟΣ ΑΝΑΦΟΡΑ", df_wrong_ref),
    ("ΔΕΝ ΒΡΕΘΗΚΑΝ", df_not_found), ("ΛΑΘΟΣ ΗΜΕΡΟΜΗΝΙΑ", df_date),
    ("ΛΑΘΟΣ ΑΞΙΑ", df_amount)
]:
    ws = wb[sheet_name]
    if ws.max_row > 1:
        tab = Table(
            displayName=sheet_name.replace(" ", "_"),
            ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        )
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

wb.save(OUTPUT)

print("Phase 2 complete — matching report generated.")
print(f"  ΒΡΕΘΗΚΑΝ:          {len(df_found)}")
print(f"  ΛΑΘΟΣ ΑΝΑΦΟΡΑ:     {len(df_wrong_ref)}")
print(f"  ΔΕΝ ΒΡΕΘΗΚΑΝ:      {len(df_not_found)}")
print(f"  ΛΑΘΟΣ ΗΜΕΡΟΜΗΝΙΑ:  {len(df_date)}")
print(f"  ΛΑΘΟΣ ΑΞΙΑ:        {len(df_amount)}")
